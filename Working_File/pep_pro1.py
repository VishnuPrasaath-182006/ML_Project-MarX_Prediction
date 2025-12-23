import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from io import BytesIO
import json
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="MarX Prediction")

# ----------------- Initialize session_state -----------------
if "page" not in st.session_state:
    st.session_state.page = "auth"
if "user_info" not in st.session_state:
    st.session_state.user_info = {}
if "num_subjects" not in st.session_state:
    st.session_state.num_subjects = 0
if "num_students" not in st.session_state:
    st.session_state.num_students = 1
if "student_data" not in st.session_state:
    st.session_state.student_data = None

# ----------------- Navigation Helper -----------------
def navigate(page_name):
    st.session_state.page = page_name



# ------------------------- Helper Functions -------------------------
def calculate_marks(study_hours, mode):
    mode_multiplier = {"Normal": 10, "Intermediate": 8, "Fast": 5}
    return round(study_hours * mode_multiplier.get(mode, 1), 2)

def generate_recommendation(study_hours):
    if study_hours > 8:
        return "Take adequate sleep"
    elif study_hours < 5:
        return "Practice meditation / yoga"
    else:
        return "Maintain regular study schedule"

def to_excel(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Prediction Summary"
    
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
    
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    
    wb.save(output)
    processed_data = output.getvalue()
    return processed_data

def to_pdf(df):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    import matplotlib.pyplot as plt
    from io import BytesIO

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("MarX Prediction Summary", styles['Title']))
    elements.append(Spacer(1, 20))

    # ---------------- Table ----------------
    table_data = [df.columns.tolist()] + df.values.tolist()
    table = Table(table_data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
    ])
    table.setStyle(style)
    elements.append(table)
    elements.append(Spacer(1, 20))

    # ---------------- Plots ----------------
    plots = []

    # 1. Bar chart: Average Marks per Subject
    marks_per_subject = df.groupby("Subject")["Predicted Marks"].mean().reset_index()
    fig, ax = plt.subplots(figsize=(6,4))
    ax.bar(marks_per_subject['Subject'], marks_per_subject['Predicted Marks'], color='skyblue')
    ax.set_title("Average Marks per Subject")
    ax.set_ylabel("Marks")
    buf = BytesIO()
    plt.tight_layout()
    fig.savefig(buf, format='png')
    buf.seek(0)
    plt.close(fig)
    plots.append(("Average Marks per Subject", buf))

    # 2. Pie chart: Total Study Hours per Subject
    study_hours_subject = df.groupby("Subject")["Study Hours"].sum().reset_index()
    fig, ax = plt.subplots(figsize=(6,4))
    ax.pie(study_hours_subject['Study Hours'], labels=study_hours_subject['Subject'], autopct='%1.1f%%')
    ax.set_title("Study Hours Distribution by Subject")
    buf2 = BytesIO()
    plt.tight_layout()
    fig.savefig(buf2, format='png')
    buf2.seek(0)
    plt.close(fig)
    plots.append(("Study Hours Distribution by Subject", buf2))

    # 3. Histogram: Marks distribution
    fig, ax = plt.subplots(figsize=(6,4))
    ax.hist(df['Predicted Marks'], bins=10, color='lightgreen', edgecolor='black')
    ax.set_title("Histogram of Marks")
    ax.set_xlabel("Marks")
    ax.set_ylabel("Frequency")
    buf3 = BytesIO()
    plt.tight_layout()
    fig.savefig(buf3, format='png')
    buf3.seek(0)
    plt.close(fig)
    plots.append(("Histogram of Marks", buf3))

    # 4. Scatter: Study Hours vs Predicted Marks per Subject
    fig, ax = plt.subplots(figsize=(6,4))
    subjects = df['Subject'].unique()
    colors_list = ['#4F81BD', '#C0504D', '#9BBB59', '#8064A2', '#F79646']
    for i, subj in enumerate(subjects):
        subset = df[df['Subject']==subj]
        ax.scatter(subset['Study Hours'], subset['Predicted Marks'], label=subj, color=colors_list[i%len(colors_list)])
    ax.set_title("Study Hours vs Predicted Marks")
    ax.set_xlabel("Study Hours")
    ax.set_ylabel("Predicted Marks")
    ax.legend()
    buf4 = BytesIO()
    plt.tight_layout()
    fig.savefig(buf4, format='png')
    buf4.seek(0)
    plt.close(fig)
    plots.append(("Study Hours vs Predicted Marks", buf4))

    # ---------------- Add plots to PDF ----------------
    for title, buf in plots:
        elements.append(Paragraph(title, styles['Heading2']))
        elements.append(Image(buf, width=400, height=250))
        elements.append(Spacer(1, 20))

    doc.build(elements)
    pdf_data = output.getvalue()
    return pdf_data


# ------------------------- Authentication -------------------------
if st.session_state.page == "auth":
    st.title("Welcome to MarX Prediction📊📈")
    st.subheader("Please Register or Sign In")
    auth_tab = st.tabs(["Register", "Sign In"])
    
    with auth_tab[0]:
        name = st.text_input("Name")
        email = st.text_input("Email")
        password = st.text_input("Password", type="password")
        if st.button("Register"):
            if name and email and password:
                st.session_state.user_info = {"name": name, "email": email}
                st.success(f"Welcome {name}! Registration successful.")
                navigate("subject_setup")
            else:
                st.error("Please fill all fields.")
                
    with auth_tab[1]:
        email_login = st.text_input("Email", key="login_email")
        password_login = st.text_input("Password", type="password", key="login_pwd")
        if st.button("Sign In"):
            if email_login and password_login:
                st.session_state.user_info = {"name": "User", "email": email_login}
                st.success(f"Welcome back! Login successful.")
                navigate("subject_setup")
            else:
                st.error("Please fill all fields.")

# ------------------------- Subject Setup -------------------------
elif st.session_state.page == "subject_setup":
    st.title("Setup Subjects for Prediction")
    st.write("Choose input method for student data.")
    input_method = st.radio("Select Input Method:", ["Single Student", "Two Students", "Upload CSV"])
    
    if input_method in ["Single Student", "Two Students"]:
        num_students = 1 if input_method == "Single Student" else 2
        num_subjects = st.number_input("Enter number of subjects:", min_value=1, max_value=20, step=1)
        if st.button("Proceed"):
            st.session_state.num_subjects = num_subjects
            st.session_state.num_students = num_students
            navigate("subject_details")
    
    elif input_method == "Upload CSV":
        uploaded_file = st.file_uploader("Upload CSV file", type=["csv"])
        if uploaded_file:
            df = pd.read_csv(uploaded_file)
            # Check if required columns exist, else raise error
            required_cols = ["Student", "Subject", "Study Hours", "Mode"]
            if not all(col in df.columns for col in required_cols):
                st.error(f"CSV must contain columns: {required_cols}")
            else:
            # Calculate predicted marks and recommendations
                df["Predicted Marks"] = df.apply(lambda row: calculate_marks(row["Study Hours"], row["Mode"]), axis=1)
                df["Recommendation"] = df["Study Hours"].apply(generate_recommendation)
                st.session_state.student_data = df
                st.success("CSV uploaded successfully!")
                if st.button("Next"):
                    navigate("recommendations")

# ------------------------- Subject Details -------------------------
elif st.session_state.page == "subject_details":
    st.title("Enter Subject Details")
    student_data = []
    modes = ["Normal", "Intermediate", "Fast"]
    
    with st.form("subject_form"):
        for student_idx in range(st.session_state.num_students):
            st.subheader(f"Student {student_idx + 1}")
            student_name = st.text_input(f"Student {student_idx + 1} Name", key=f"name_{student_idx}")
            subjects = []
            for subj_idx in range(st.session_state.num_subjects):
                col1, col2, col3 = st.columns([3,3,3])
                with col1:
                    subject_name = st.text_input(f"Subject {subj_idx + 1} Name", key=f"subj_name_{student_idx}_{subj_idx}")
                with col2:
                    study_hours = st.number_input(f"Study Hours", min_value=0.0, max_value=24.0, step=0.5, key=f"hours_{student_idx}_{subj_idx}")
                with col3:
                    mode = st.selectbox(f"Study Mode", modes, key=f"mode_{student_idx}_{subj_idx}")
                predicted_marks = calculate_marks(study_hours, mode)
                recommendation = generate_recommendation(study_hours)
                subjects.append({"Student": student_name, "Subject": subject_name, "Study Hours": study_hours,
                                 "Mode": mode, "Predicted Marks": predicted_marks, "Recommendation": recommendation})
            student_data.extend(subjects)
        submitted = st.form_submit_button("Submit")
        if submitted:
            st.session_state.student_data = pd.DataFrame(student_data)
            navigate("recommendations")

# ------------------------- Recommendations & Visuals -------------------------
elif st.session_state.page == "recommendations":
    df = st.session_state.student_data
    st.title("Predicted Marks & Recommendations")
    
    st.subheader("Summary Table")
    st.dataframe(df.style.set_properties(**{'background-color': '#f9f9f9',
                                           'color': 'black',
                                           'border-color': 'black'}))
    
    st.subheader("Visualizations")
    col1, col2 = st.columns(2)
    
    marks_per_subject = df.groupby("Subject")["Predicted Marks"].mean().reset_index()
    mode_distribution = df.groupby("Mode")["Study Hours"].sum().reset_index()
    
    with col1:
        plt.figure(figsize=(6,4))
        plt.bar(marks_per_subject['Subject'], marks_per_subject['Predicted Marks'], color='skyblue')
        plt.title("Average Marks per Subject")
        plt.ylabel("Marks")
        st.pyplot(plt)
        plt.close()
        
        plt.figure(figsize=(6,4))
        plt.pie(mode_distribution['Study Hours'], labels=mode_distribution['Mode'], autopct='%1.1f%%', colors=['#4F81BD','#C0504D','#9BBB59'])
        plt.title("Study Hours Distribution by Mode")
        st.pyplot(plt)
        plt.close()
    
    with col2:
        plt.figure(figsize=(6,4))
        plt.hist(df['Predicted Marks'], bins=10, color='lightgreen', edgecolor='black')
        plt.title("Histogram of Marks")
        plt.xlabel("Marks")
        plt.ylabel("Frequency")
        st.pyplot(plt)
        plt.close()
        
        plt.figure(figsize=(6,4))
        for mode in df['Mode'].unique():
            subset = df[df['Mode']==mode]
            plt.scatter(subset['Subject'], subset['Predicted Marks'], label=mode)
        plt.title("Marks per Subject by Mode")
        plt.xlabel("Subject")
        plt.ylabel("Predicted Marks")
        plt.legend()
        st.pyplot(plt)
        plt.close()
    
    st.subheader("Download Data")
    st.download_button("Download Excel", to_excel(df), file_name="prediction.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.download_button("Download JSON", df.to_json(orient="records", indent=4), file_name="prediction.json", mime="application/json")
    st.download_button("Download PDF", to_pdf(df), file_name="prediction.pdf", mime="application/pdf")
    
    st.success("Prediction Complete! Scroll up to view summary and charts.")
